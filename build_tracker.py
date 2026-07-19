"""Build target_cost_tracker.xlsx — live-formula Excel twin of should_cost.py.
Blue = hardcoded inputs · black = formulas · yellow fill = key levers to edit.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial")
BOLD = Font(name="Arial", bold=True)
HDR = Font(name="Arial", bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="1F4E79")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Border(bottom=Side(style="thin", color="B0B0B0"))
YEN = '"¥"#,##0'
YEN1 = '"¥"#,##0.0'
PCT = '0.0%'

wb = Workbook()

# ---------------- Assumptions ----------------
ws = wb.active; ws.title = "Assumptions"
ws["A1"] = "Assumptions & rates (synthetic, illustrative)"; ws["A1"].font = Font(name="Arial", bold=True, size=13)
ws["A2"] = "Blue cells = inputs you can edit; yellow = key levers. All costs JPY."; ws["A2"].font = Font(name="Arial", italic=True, size=9)

globals_ = [("Scrap %", 0.03), ("Overhead %", 0.12), ("Supplier margin %", 0.10),
            ("Catalog handling %", 0.02), ("Target reduction %", 0.08), ("Annual volume (units)", 120000)]
ws["A4"] = "Global levers"; ws["A4"].font = BOLD
for i,(k,v) in enumerate(globals_, start=5):
    ws.cell(row=i, column=1, value=k).font = BLACK
    c = ws.cell(row=i, column=2, value=v); c.font = BLUE; c.fill = YELLOW
    c.number_format = PCT if "%" in k else '#,##0'

ws["D4"] = "Material rates (JPY/kg, drawing parts)"; ws["D4"].font = BOLD
MAT = [("FR4",2400),("Alloy",3200),("Aluminum",950),("Ceramic",3000),("PA66",780),
       ("Ferrite",1500),("Steel",320),("Copper",1450),("SUS304",680),("ADC12",640),
       ("PPS",1250),("Carbon/SiC",8000),("EPDM",900),("Silicone",800),("PET",400),("PE",260)]
for i,(m,r) in enumerate(MAT, start=5):
    ws.cell(row=i, column=4, value=m).font = BLACK
    c = ws.cell(row=i, column=5, value=r); c.font = BLUE; c.number_format = '#,##0'

ws["G4"] = "Process rates by commodity"; ws["G4"].font = BOLD
ws["G5"] = "Commodity"; ws["H5"] = "Var (JPY/g)"; ws["I5"] = "Fixed (JPY)"
for cell in ("G5","H5","I5"): ws[cell].font = BOLD
PROC = [("Electromech",0.55,40),("Mechanical",0.5,18),("Fasteners",0.1,4),("Packaging",0.1,3)]
for i,(c_,v,f) in enumerate(PROC, start=6):
    ws.cell(row=i, column=7, value=c_).font = BLACK
    a = ws.cell(row=i, column=8, value=v); a.font = BLUE
    b = ws.cell(row=i, column=9, value=f); b.font = BLUE
ws["G11"] = "Catalog parts (Electronics) are benchmarked at market_ref x (1 + catalog handling), not weight-costed."
ws["G11"].font = Font(name="Arial", italic=True, size=9)
for col,w in zip("ABCDEFGHI", (26,12,3,14,12,3,14,11,11)): ws.column_dimensions[col].width = w

# ---------------- BOM ----------------
bom = wb.create_sheet("BOM")
rows = list(csv.DictReader(open("data/bom.csv")))
headers = ["part_no","part_name","commodity","supplier","material","qty","weight_g",
           "current_price_jpy","market_ref_jpy","current_ext_jpy","should_cost_unit_jpy",
           "should_cost_ext_jpy","gap_jpy","annual_gap_jpy"]
for j,h in enumerate(headers,1):
    c = bom.cell(row=1, column=j, value=h); c.font = HDR; c.fill = HDRFILL
for i,r in enumerate(rows, start=2):
    vals = [r["part_no"], r["part_name"], r["commodity"], r["supplier"], r["material"],
            int(r["qty_per_unit"]), float(r["unit_weight_g"]), float(r["current_price_jpy"]),
            float(r["market_ref_jpy"]) if r["market_ref_jpy"] else None]
    for j,v in enumerate(vals,1):
        c = bom.cell(row=i, column=j, value=v)
        c.font = BLUE if j >= 6 else BLACK
        c.border = THIN
    bom.cell(row=i, column=10, value=f"=F{i}*H{i}").number_format = YEN                    # current ext
    bom.cell(row=i, column=11, value=(                                                     # should unit
        f'=IF(C{i}="Electronics",I{i}*(1+Assumptions!$B$8),'
        f'((G{i}/1000)*INDEX(Assumptions!$E$5:$E$20,MATCH(E{i},Assumptions!$D$5:$D$20,0))*(1+Assumptions!$B$5)'
        f'+G{i}*INDEX(Assumptions!$H$6:$H$9,MATCH(C{i},Assumptions!$G$6:$G$9,0))'
        f'+INDEX(Assumptions!$I$6:$I$9,MATCH(C{i},Assumptions!$G$6:$G$9,0)))'
        f'*(1+Assumptions!$B$6)*(1+Assumptions!$B$7))')).number_format = YEN1
    bom.cell(row=i, column=12, value=f"=K{i}*F{i}").number_format = YEN1                   # should ext
    bom.cell(row=i, column=13, value=f"=J{i}-L{i}").number_format = YEN1                   # gap
    bom.cell(row=i, column=14, value=f"=M{i}*Assumptions!$B$10").number_format = YEN       # annual gap
    for j in range(10,15): bom.cell(row=i, column=j).font = BLACK; bom.cell(row=i, column=j).border = THIN
last = len(rows) + 1
widths = (9,30,13,18,11,6,9,15,14,14,17,16,11,15)
for col,w in zip("ABCDEFGHIJKLMN", widths): bom.column_dimensions[col].width = w
bom.freeze_panes = "A2"

# ---------------- Summary ----------------
s = wb.create_sheet("Summary")
s["A1"] = "Target-cost attainment — electric coolant pump assembly (synthetic BOM)"
s["A1"].font = Font(name="Arial", bold=True, size=13)
s["A2"] = "All figures recalculate from BOM + Assumptions. Data is illustrative; the method is the deliverable."
s["A2"].font = Font(name="Arial", italic=True, size=9)
items = [
    ("Current assembly cost", f"=SUM(BOM!J2:J{last})", YEN),
    ("Should-cost (bottom-up + market benchmark)", f"=SUM(BOM!L2:L{last})", YEN),
    ("Target cost", "=B4*(1-Assumptions!$B$9)", YEN),
    ("Gap: current vs should-cost", "=B4-B5", YEN),
    ("Gap % of current", "=B7/B4", PCT),
    ("Required reduction to target", "=B4-B6", YEN),
    ("Target attainment (should-cost achieved)", "=B7/B9", PCT),
    ("Annual value @ volume", "=B7*Assumptions!$B$10", YEN),
]
for i,(k,f,fmt) in enumerate(items, start=4):
    s.cell(row=i, column=1, value=k).font = BLACK
    c = s.cell(row=i, column=2, value=f); c.font = BOLD; c.number_format = fmt
s["A13"] = "Gap by commodity"; s["A13"].font = BOLD
s["A14"] = "Commodity"; s["B14"] = "Current"; s["C14"] = "Should"; s["D14"] = "Gap"
for cell in ("A14","B14","C14","D14"): s[cell].font = HDR; s[cell].fill = HDRFILL
for i,c_ in enumerate(["Electromech","Mechanical","Electronics","Fasteners","Packaging"], start=15):
    s.cell(row=i, column=1, value=c_).font = BLACK
    s.cell(row=i, column=2, value=f'=SUMIFS(BOM!$J$2:$J${last},BOM!$C$2:$C${last},A{i})').number_format = YEN
    s.cell(row=i, column=3, value=f'=SUMIFS(BOM!$L$2:$L${last},BOM!$C$2:$C${last},A{i})').number_format = YEN
    s.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = YEN
    for j in (2,3,4): s.cell(row=i, column=j).font = BLACK
s["A21"] = "How to use: edit blue cells (Assumptions levers, BOM prices/weights) — everything recalculates."
s["A21"].font = Font(name="Arial", italic=True, size=9)
s["A22"] = "Example: raise Copper to 1,600 JPY/kg to simulate a commodity shock; attainment updates instantly."
s["A22"].font = Font(name="Arial", italic=True, size=9)
for col,w in zip("ABCD",(42,16,16,14)): s.column_dimensions[col].width = w

wb.save("target_cost_tracker.xlsx")
print("workbook written")
